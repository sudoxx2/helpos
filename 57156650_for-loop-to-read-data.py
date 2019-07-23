from openpyxl import load_workbook

# get email from each row based on parameters passed
def email_id(row_Num, cell_Num, sheet):
    email = sheet.cell(row=row_Num, column=cell_Num).value

    if email is not None:
        return email
    else:
        return ''

# setup our sheet to be easily passed through various functions
def get_sheet():
    wb = load_workbook('/Users/sudoxx2/Downloads/test_sheet.xlsx')
    sheet = wb.active
    return sheet

# main entry point to our python script(universal entry point in most python scripts)
if __name__ == "__main__":
    curr_sheet = get_sheet()
    for i in range(1, curr_sheet.max_row + 1):
        print(email_id(i, 1, curr_sheet))
